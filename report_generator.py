# -*- coding: utf-8 -*-
"""
台股本益比評價與價值投資選股系統 - Excel 報表生成模組
負責將分析完畢的台股數據，利用 openpyxl 導出為具備「投行法人質感」的 Excel 報表。
【美化設計與專業排版】：
1. 標題欄位採用「海軍藍/深鋼藍 (Deep Steel Blue)」背景配白色粗體字。
2. 資料儲存格自動適應列寬 (Auto-fit Columns)，防止中文字被截斷或出現 ### 亂碼。
3. 專業數值格式化：金額加千分位 (e.g. 1,234.50)，百分比格式化 (e.g. 12.50%)。
4. 智能著色：財務體質優良的個股以「淡莫蘭迪綠」高亮；低估率大於 20% 的黃金股以「淡橘色」高亮，利於決策。
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_PATH

class ExcelReportGenerator:
    """
    Excel 報表美化與輸出類別
    """
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        
    def export_to_excel(self):
        """
        導出數據並對 Excel 檔案進行全面的字型、顏色、對齊與格式設定。
        """
        print(f" Excel 正在導出並美化報表，準備寫入至：{os.path.abspath(EXCEL_PATH)}...")
        
        # 1. 準備輸出的欄位映射與順序 (繁體中文標題，讓初學者一目了然)
        column_mapping = {
            "rank": "綜合排名",
            "symbol": "股票代號",
            "name": "股票名稱",
            "close_price": "最新收盤價",
            "latest_eps": "最新年度EPS",
            "pe": "目前本益比(P/E)",
            "pe_avg_5y": "5年平均本益比",
            "pe_avg_10y": "10年平均本益比",
            "fair_price": "合理價",
            "cheap_price": "便宜價(買點)",
            "expensive_price": "昂貴價(賣點)",
            "undervalued_rate": "低估幅度(%)",
            "pb_ratio": "股價淨值比(P/B)",
            "dividend_yield": "股息殖利率",
            "roe": "股東權益報酬率(ROE)",
            "debt_ratio": "負債比率(%)",
            "peg": "PEG比率",
            "peg_status": "成長性評估",
            "f_score": "F-Score體質分",
            "dcf_valuation": "DCF內在價值",
            "gordon_valuation": "高登估值",
            "value_score": "法人價值分數",
            "health_status": "財務體質狀態",
            "cyclical_sector": "景氣循環分類"
        }
        
        # 新增一個排名欄位 (1-based)
        export_df = self.df.copy()
        export_df.insert(0, "rank", range(1, len(export_df) + 1))
        
        # 過濾並重命名欄位
        export_df = export_df[list(column_mapping.keys())]
        export_df = export_df.rename(columns=column_mapping)
        
        # 2. 建立 Excel 活頁簿
        wb = Workbook()
        ws = wb.active
        ws.title = "台股價值選股排行榜"
        
        # 確保 Excel 顯示格線
        ws.views.sheetView[0].showGridLines = True
        
        # 3. 定義專業投行樣式與調色盤
        font_name = "Microsoft JhengHei" # 微軟正黑體
        
        # 字型
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_name, size=10, color="000000")
        bold_data_font = Font(name=font_name, size=10, bold=True, color="000000")
        
        # 填充顏色 (莫蘭迪色系)
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")     # 海軍藍
        healthy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")    # 淡雅綠 (財務良好)
        undervalued_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")# 淡橘色 (極度低估)
        cyclical_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # 輕盈灰 (景氣循環股)
        
        # 對齊方式
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 框線
        thin_border_side = Side(style="thin", color="D3D3D3")
        border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # 4. 寫入標題列並套用樣式
        headers = list(export_df.columns)
        ws.append(headers)
        ws.row_dimensions[1].height = 28 # 加高標題列
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        # 5. 寫入資料列並逐一設定格式與高亮
        for row_num, row_data in enumerate(export_df.values, 2):
            ws.append(list(row_data))
            ws.row_dimensions[row_num].height = 22 # 適度加寬行高，閱讀更舒適
            
            # 獲取本行關鍵欄位值以決定是否著色
            # 股票代號位置: Column 2 | 體質狀態位置: Column 23 | 低估率位置: Column 12 | 循環股位置: Column 24
            symbol_val = ws.cell(row=row_num, column=2).value
            undervalued_val = ws.cell(row=row_num, column=12).value
            health_status_val = ws.cell(row=row_num, column=23).value
            cyclical_val = ws.cell(row=row_num, column=24).value
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = border_all
                
                # A. 數據對齊設定
                if col_num in [1, 2, 18, 19]: # 排名、代號、成長評估、F-Score
                    cell.alignment = align_center
                elif col_num in [3, 23, 24]:  # 名稱、體質、循環股別
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    
                # B. 數值與格式化設定
                val = cell.value
                if val is not None and not isinstance(val, str):
                    # 收盤價、合理價、便宜、昂貴、DCF、Gordon 價格格式
                    if col_num in [4, 9, 10, 11, 20, 21]:
                        cell.number_format = "#,##0.00"
                    # EPS
                    elif col_num in [5]:
                        cell.number_format = "0.00"
                    # PE, 5yPE, 10yPE, PB, PEG, 價值分數
                    elif col_num in [6, 7, 8, 13, 17, 22]:
                        cell.number_format = "0.00"
                    # 低估率、ROE、負債比
                    elif col_num in [12, 15, 16]:
                        cell.value = val / 100.0 # 存入 Excel 時轉為小數，以便搭配 % 顯示
                        cell.number_format = "0.0%"
                    # 股息殖利率 (原本已是小數)
                    elif col_num in [14]:
                        cell.number_format = "0.00%"
                        
                # C. 條件格式化智能著色 (莫蘭迪色系，低飽和度不刺眼)
                # 情況一：如果低估率 > 20%，將低估幅度欄位標示為淡橙色
                if col_num == 12 and undervalued_val is not None and undervalued_val > 20:
                    cell.fill = undervalued_fill
                    cell.font = bold_data_font
                    
                # 情況二：如果財務體質優良，將整個財務狀態儲存格標示為淡雅莫蘭迪綠
                if col_num == 23 and health_status_val == "財務體質優良":
                    cell.fill = healthy_fill
                    cell.font = bold_data_font
                    
                # 情況三：如果是景氣循環股，將其該列代號、名稱與循環標記高亮為淡灰色，作為警示
                if cyclical_val != "非景氣循環股" and col_num in [2, 3, 24]:
                    cell.fill = cyclical_fill
                    
        # 6. 自動調適所有直欄的寬度 (Auto-fit Column Widths)
        # 遍歷每一列，找出最大長度，並加上 padding 設定寬度
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                val_str = str(cell.value or "")
                # 處理中文字元長度 (中文字在 Excel 佔 2 個字元寬)
                byte_len = 0
                for char in val_str:
                    if ord(char) > 127:
                        byte_len += 2
                    else:
                        byte_len += 1
                max_len = max(max_len, byte_len)
                
            # 設定寬度，最小給 12，最大不超過 25，防止極長備註欄撐破版面
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 25)
            
        # 7. 保存檔案
        # 確保資料夾存在
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        wb.save(EXCEL_PATH)
        print("🎉 Excel 報表美化輸出完成！")

# 單獨測試此模組
if __name__ == "__main__":
    from data_fetcher import StockDataFetcher
    from analyzer import StockAnalyzer
    fetcher = StockDataFetcher()
    from config import STOCK_POOL
    df_raw = fetcher.fetch_all_data(STOCK_POOL)
    analyzer = StockAnalyzer(df_raw)
    df_analyzed = analyzer.run_analysis()
    
    generator = ExcelReportGenerator(df_analyzed)
    generator.export_to_excel()
